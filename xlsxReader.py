import openpyxl
import os
import matplotlib.pyplot as plt
import math
import xlsxwriter
from PIL import Image
import shutil
import time
print("Starting...")
file = "./Output - Backup.xlsx"
wb = openpyxl.load_workbook(file, read_only=True)
sheet = wb.active
sheet.iter_rows()
m_row = sheet.max_row
m_col = sheet.max_column
img_path = './media'
imgs = next(os.walk(img_path))[2]
imgs.sort(key=lambda x: int(''.join(filter(str.isdigit, x))))# Taken from https://stackoverflow.com/questions/36259763/sort-list-of-string-based-on-number-in-string
im = []
count = 3
for img in imgs:
    name = img.replace("image","").replace(".png","")
    num = str(count)
    if num == name:
        im.append(img_path+"/"+img)
        '''with Image.open(img_path+"/"+img) as showImg:
            showImg.show()
            showImg.close()'''
        count += 4
#print(sheet.cell(row = 2, column = 1).value)
global MassPct
MassPct = []


def pprint(x:str):
    for items in x:
        print(items,end='')
    print("",end='\n')

def resize(path, basewidth):
    #basewidth = 300
    if not os.path.exists('./xlsxCache/'):
        os.makedirs('./xlsxCache/')
    img = Image.open(path)
    wpercent = (basewidth/float(img.size[0]))
    hsize = int((float(img.size[1])*float(wpercent)))
    img = img.resize((basewidth,hsize), Image.LANCZOS)
    #imgFormat = path.split("/")[0]
    imgFormat = path.split(".")[-1]
    imgPath = './xlsxCache/' + str(time.time()).replace(".","") +"."+ imgFormat
    img.save(imgPath)
    return imgPath

def clearCache():
    if os.path.exists('./xlsxCache/'):
        shutil.rmtree('./xlsxCache')

def avg(*args):
    #Returns Average of all args
    return sum(args)/len(args)

def ploting(image,y=[],x=[],save=None):
    fig, ax = plt.subplots()
    ax.imshow(image, extent=[0, 223.5, 223.5, 0])
    for line in x:
        plt.axvline(x = line, color = 'b', linestyle = '-')
    for line in y:
        plt.axhline(y = line, color = 'b', linestyle = '-')
    if save:
        plt.savefig(save)
        plt.close()
    else:
        plt.show()
        #pass
    return save
    


def Calculate(image, name, data, index):
    pprint("Calculating "+name)
    global MassPct
    global worksheet
    global data_format
    GoldenRatio = 1.618033988749895
    d = [None]#for simplicity, index 0 is empty
    x = 'x'   #for convinience
    y = 'y'
    for items in data:
        d.append(items)
    graphs = []
    #variables below are in lists, index 0 - index 1 is the longer distance(Quotion), index 2 - index 3 is the shorter distance (Divisor)
    v1 = [avg(d[30][y],d[31][y]),avg(d[37][y], d[40][y], d[43][y], d[46][y]),max(d[32][y],d[36][y]),avg(d[30][y],d[31][y])]# 1. (Eyes to Nose Flair) to Nose Base
    graphs.append(ploting(image,v1,save="./xlsxReaderCache/"+str(time.time()).replace(".","")+".png"))
    v2 = [avg(d[31][y],d[34][y]),avg(d[37][y], d[40][y], d[43][y], d[46][y]),avg(d[62][y],d[63][y],d[64][y],d[66][y],d[67][y],d[68][y]),avg(d[31][y],d[34][y])]# 2. (Eyes to Nostril top) to Centre of lips
    graphs.append(ploting(image,v2,save="./xlsxReaderCache/"+str(time.time()).replace(".","")+".png"))
    v3 = [max(d[32][y],d[33][y],d[34][y],d[35][y],d[36][y]),avg(d[37][y], d[40][y], d[43][y], d[46][y]),max(d[56][y],d[57][y],d[58][y],d[59][y],d[60][y]),max(d[32][y],d[33][y],d[34][y],d[35][y],d[36][y])]# 3. (Eyes to Nose base) to Bottom of lips
    graphs.append(ploting(image,v3,save="./xlsxReaderCache/"+str(time.time()).replace(".","")+".png"))
    v4 = [avg(d[62][y],d[63][y],d[64][y],d[66][y],d[67][y],d[68][y]),avg(d[37][y], d[40][y], d[43][y], d[46][y]),max(d[7][y],d[8][y],d[9][y],d[10][y],d[11][y]),avg(d[62][y],d[63][y],d[64][y],d[66][y],d[67][y],d[68][y])]# 4. (Eyes to Centre of lips) to Bottom of Chin
    graphs.append(ploting(image,v4,save="./xlsxReaderCache/"+str(time.time()).replace(".","")+".png"))
    v5 = [max(d[56][y],d[57][y],d[58][y],d[59][y],d[60][y]),avg(d[30][y],d[31][y]),max(d[56][y],d[57][y],d[58][y],d[59][y],d[60][y]),max(d[7][y],d[8][y],d[9][y],d[10][y],d[11][y])]# 5. (Nose Flair to Bottom of Lips) to Bottom of Chin
    graphs.append(ploting(image,v5,save="./xlsxReaderCache/"+str(time.time()).replace(".","")+".png"))
    v6 = [max(d[7][y],d[8][y],d[9][y],d[10][y],d[11][y]),min(d[50][y],d[51][y],d[52][y],d[53][y],d[54][y]),avg(d[30][y],d[31][y]),min(d[50][y],d[51][y],d[52][y],d[53][y],d[54][y])]# 6. (Nose Flair to Top of Lips) to Bottom of Chin
    graphs.append(ploting(image,v6,save="./xlsxReaderCache/"+str(time.time()).replace(".","")+".png"))
    v7 = [max(d[7][y],d[8][y],d[9][y],d[10][y],d[11][y]),max(d[56][y],d[57][y],d[58][y],d[59][y],d[60][y]),max(d[56][y],d[57][y],d[58][y],d[59][y],d[60][y]),min(d[50][y],d[51][y],d[52][y],d[53][y],d[54][y])]# 7. (Top of Lips to Bottom of lips) to Bottom of Chin
    graphs.append(ploting(image,v7,save="./xlsxReaderCache/"+str(time.time()).replace(".","")+".png"))
    v8 = [max(d[56][y],d[57][y],d[58][y],d[58][y],d[59][y],d[60][y]),avg(d[62][y],d[63][y],d[64][y],d[66][y],d[67][y],d[68][y]),avg(d[62][y],d[63][y],d[64][y],d[66][y],d[67][y],d[68][y]),min(d[50][y],d[51][y],d[52][y],d[53][y],d[54][y])]# 8. (Top of Lips to Centre of lips) to Bottom of Lips
    graphs.append(ploting(image,v8,save="./xlsxReaderCache/"+str(time.time()).replace(".","")+".png"))
    v9 = [min(d[38][y],d[39][y],d[44][y],d[45][y]),min(d[18][y],d[19][y],d[20][y],d[21][y],d[22][y],d[23][y],d[24][y],d[25][y],d[26][y],d[27][y]),max(d[41][y],d[42][y],d[47][y],d[48][y]),min(d[38][y],d[39][y],d[44][y],d[45][y])]# 9. (Top of Eyebrows to Top of eyes) to Bottom of eyes
    graphs.append(ploting(image,v9,save="./xlsxReaderCache/"+str(time.time()).replace(".","")+".png"))
    v10 = [min(d[50][y],d[51][y],d[52][y],d[53][y],d[54][y]),min(d[18][y],d[19][y],d[20][y],d[21][y],d[22][y],d[23][y],d[24][y],d[25][y],d[26][y],d[27][y]),max(d[7][y],d[8][y],d[9][y],d[10][y],d[11][y]),min(d[50][y],d[51][y],d[52][y],d[53][y],d[54][y])]# 10. (Top of Eyebrows to Top of lips) to Bottom of Chin
    graphs.append(ploting(image,v10,save="./xlsxReaderCache/"+str(time.time()).replace(".","")+".png"))
    v11a = [d[37][x],min(d[1][x],d[2][x],d[3][x],d[4][x]),avg(d[38][x],d[39][x],d[41][x],d[42][x]),d[37][x]]# 11a. (Left Side of Face to Left Side of eyes) to Centre of Left Pupil
    graphs.append(ploting(image,x=v11a,save="./xlsxReaderCache/"+str(time.time()).replace(".","")+".png"))
    v11b = [d[46][x],max(d[14][x],d[15][x],d[16][x],d[17][x]),avg(d[44][x],d[45][x],d[47][x],d[48][x]),d[46][x]]# 11b. (Right Side of Face to Right Side of eyes) to Centre of Right Pupil
    graphs.append(ploting(image,x=v11b,save="./xlsxReaderCache/"+str(time.time()).replace(".","")+".png"))
    v12a = [avg(d[38][x],d[42][x]),min(d[1][x],d[2][x],d[3][x],d[4][x]),d[40][x],avg(d[38][x],d[42][x])]# 12a. (Left Side of Face to Left Side of Iris) to Other Side of Left eye
    graphs.append(ploting(image,x=v12a,save="./xlsxReaderCache/"+str(time.time()).replace(".","")+".png"))
    v12b = [avg(d[45][x],d[47][x]),max(d[14][x],d[15][x],d[16][x],d[17][x]),d[43][x],avg(d[45][x],d[47][x])]# 12b. (Right Side of Face to Right Side of Iris) to Other Side of Right eye
    graphs.append(ploting(image,x=v12b,save="./xlsxReaderCache/"+str(time.time()).replace(".","")+".png"))
    v13a = [avg(d[39][x],d[41][x]),min(d[1][x],d[2][x],d[3][x],d[4][x]),avg(min(d[1][x],d[2][x],d[3][x],d[4][x]),max(d[14][x],d[15][x],d[16][x],d[17][x])),avg(d[39][x],d[41][x])]# 13a. (Left Side of Face to Left Side of Iris) to Centre of Face
    graphs.append(ploting(image,x=v13a,save="./xlsxReaderCache/"+str(time.time()).replace(".","")+".png"))
    v13b = [avg(d[45][x],d[47][x]),max(d[14][x],d[15][x],d[16][x],d[17][x]),avg(min(d[1][x],d[2][x],d[3][x],d[4][x]),max(d[14][x],d[15][x],d[16][x],d[17][x])),avg(d[45][x],d[47][x])]# 13b. (Right Side of Face to Right Side of Iris) to Centre of Face
    graphs.append(ploting(image,x=v13b,save="./xlsxReaderCache/"+str(time.time()).replace(".","")+".png"))
    v14a = [d[40][x],min(d[1][x],d[2][x],d[3][x],d[4][x]),d[43][x],d[40][x]]# 14a. (Left Side of Face to Side of Left eye) to the Right eye
    graphs.append(ploting(image,x=v14a,save="./xlsxReaderCache/"+str(time.time()).replace(".","")+".png"))
    v14b = [max(d[14][x],d[15][x],d[16][x],d[17][x]),d[43][x],d[43][x],d[40][x]]# 14b. (Right Side of Face to Side of Right eye) to the Left eye
    graphs.append(ploting(image,x=v14b,save="./xlsxReaderCache/"+str(time.time()).replace(".","")+".png"))
    v15a = [avg(min(d[1][x],d[2][x],d[3][x],d[4][x]),max(d[14][x],d[15][x],d[16][x],d[17][x])),min(d[1][x],d[2][x],d[3][x],d[4][x]),d[46][x],avg(min(d[1][x],d[2][x],d[3][x],d[4][x]),max(d[14][x],d[15][x],d[16][x],d[17][x]))]# 15a. (Left Side of Face to Centre of Face) to the Right eye
    graphs.append(ploting(image,x=v15a,save="./xlsxReaderCache/"+str(time.time()).replace(".","")+".png"))
    v15b = [max(d[14][x],d[15][x],d[16][x],d[17][x]),avg(min(d[1][x],d[2][x],d[3][x],d[4][x]),max(d[14][x],d[15][x],d[16][x],d[17][x])),avg(min(d[1][x],d[2][x],d[3][x],d[4][x]),max(d[14][x],d[15][x],d[16][x],d[17][x])),d[37][x]]# 15b. (Right Side of Face to Centre of Face) to the Left eye
    graphs.append(ploting(image,x=v15b,save="./xlsxReaderCache/"+str(time.time()).replace(".","")+".png"))
    v16a = [d[43][x],min(d[1][x],d[2][x],d[3][x],d[4][x]),max(d[14][x],d[15][x],d[16][x],d[17][x]),d[43][x]]# 16a. (Left Side of Face to Inside Right eye) to Right side of face
    graphs.append(ploting(image,x=v16a,save="./xlsxReaderCache/"+str(time.time()).replace(".","")+".png"))
    v16b = [max(d[14][x],d[15][x],d[16][x],d[17][x]),d[40][x],d[40][x],min(d[1][x],d[2][x],d[3][x],d[4][x])]# 16b. (Right Side of Face to Inside Left eye) to Left side of face
    graphs.append(ploting(image,x=v16b,save="./xlsxReaderCache/"+str(time.time()).replace(".","")+".png"))
    v17a = [avg(d[32][x],d[33][x]),d[37][x],avg(d[35][x],d[36][x]),avg(d[32][x],d[33][x])]# 17a. (Left Side of Eye to Left Flair of Nose) to Right Flair of Nose
    graphs.append(ploting(image,x=v17a,save="./xlsxReaderCache/"+str(time.time()).replace(".","")+".png"))
    v17b = [d[46][x],avg(d[35][x],d[36][x]),avg(d[35][x],d[36][x]),avg(d[32][x],d[33][x])]# 17b. (Right Side of Eye to Right Flair of Nose) to Left Flair of Nose
    graphs.append(ploting(image,x=v17b,save="./xlsxReaderCache/"+str(time.time()).replace(".","")+".png"))
    v18a = [avg(d[30][y],d[31][y]),avg(d[38][y],d[39][y],d[41][y],d[42][y]),max(d[32][y],d[33][y],d[34][y],d[35][y],d[36][y]),avg(d[30][y],d[31][y])]# 18a. (Centre of Left Pupil to Top of Nose Flair) to bottom of nose
    graphs.append(ploting(image,v18a,save="./xlsxReaderCache/"+str(time.time()).replace(".","")+".png"))
    v18b = [avg(d[30][y],d[31][y]),avg(d[44][y],d[45][y],d[47][y],d[48][y]),max(d[32][y],d[33][y],d[34][y],d[35][y],d[36][y]),avg(d[30][y],d[31][y])]# 18b. (Centre of Right Pupil to Top of Nose Flair) to bottom of nose
    graphs.append(ploting(image,v18b,save="./xlsxReaderCache/"+str(time.time()).replace(".","")+".png"))
    v19a = [d[33][x],avg(d[38][x],d[39][x],d[41][x],d[42][x]),d[35][x],d[36][x]]# 19a. (Centre of Left Pupil to Centre of Left Nostril) to Centre of Right Nostril
    graphs.append(ploting(image,x=v19a,save="./xlsxReaderCache/"+str(time.time()).replace(".","")+".png"))
    v19b = [avg(d[44][x],d[45][x],d[47][x],d[48][x]),d[35][x],d[35][x],d[33][x]]# 19b. (Centre of Right Pupil to Centre of Right Nostrils) to Centre of Left Nostril
    graphs.append(ploting(image,x=v19b,save="./xlsxReaderCache/"+str(time.time()).replace(".","")+".png"))
    v20a = [avg(d[33][x],d[34][x]),d[40][x],avg(d[34][x],d[35][x]),avg(d[33][x],d[34][x])]# 20a. (Inside of Left Eye to Left Nose Bridge) to Right Nose Bridge
    graphs.append(ploting(image,x=v20a,save="./xlsxReaderCache/"+str(time.time()).replace(".","")+".png"))
    v20b = [d[43][x],avg(d[34][x],d[35][x]),avg(d[34][x],d[35][x]),avg(d[33][x],d[34][x])]# 20b. (Inside of Right Eye to Right Nose Bridge) to Left Nose Bridge
    graphs.append(ploting(image,x=v20b,save="./xlsxReaderCache/"+str(time.time()).replace(".","")+".png"))
    v21a = [d[51][x],min(d[49][x],d[61][x]),d[53][x],d[51][x]]# 21a. (Left Side of Mouth to Cupid’s Left bow point of lips) to Cupid’s Right bow point of lips
    graphs.append(ploting(image,x=v21a,save="./xlsxReaderCache/"+str(time.time()).replace(".","")+".png"))
    v21b = [max(d[55][x],d[65][x]),d[53][x],d[53][x],d[51][x]]# 21b. (Right Side of Mouth to Cupid’s Right bow point of lips) to Cupid’s Left bow point of lips
    graphs.append(ploting(image,x=v21b,save="./xlsxReaderCache/"+str(time.time()).replace(".","")+".png"))
    v22 = [max(d[56][y],d[57][y],d[58][y],d[59][y],d[60][y]),min(d[51][y],d[52][y],d[53][y]),min(d[51][y],d[52][y],d[53][y]),max(d[32][y],d[33][y],d[34][y],d[35][y],d[36][y])]# 22. (Bottom of nose to Top of lips) to Bottom of lips
    graphs.append(ploting(image,v22,save="./xlsxReaderCache/"+str(time.time()).replace(".","")+".png"))
    v23 = [max(d[7][y],d[8][y],d[9][y],d[10][y],d[11][y]),d[69][y],max(d[14][x],d[15][x],d[16][x],d[17][x]),min(d[1][x],d[2][x],d[3][x],d[4][x])]# 23. Head Height to Head Width
    graphs.append(ploting(image,[v23[0],v23[1]],[v23[2],v23[3]],save="./xlsxReaderCache/"+str(time.time()).replace(".","")+".png"))
    variables = [v1,v2,v3,v4,v5,v6,v7,v8,v9,v10,v11a,v11b,v12a,v12b,v13a,v13b,v14a,v14b,v15a,v15b,v16a,v16b,v17a,v17b,v18a,v18b,v19a,v19b,v20a,v20b,v21a,v21b,v22,v23]
    varNames = ['1. (Eyes to Nose Flair) to Nose Base', '2. (Eyes to Nostril top) to Centre of lips', '3. (Eyes to Nose base) to Bottom of lips', '4. (Eyes to Centre of lips) to Bottom of Chin', '5. (Nose Flair to Bottom of Lips) to Bottom of Chin', '6. (Nose Flair to Top of Lips) to Bottom of Chin', '7. (Top of Lips to Bottom of lips) to Bottom of Chin', '8. (Top of Lips to Centre of lips) to Bottom of Lips', '9. (Top of Eyebrows to Top of eyes) to Bottom of eyes', '10. (Top of Eyebrows to Top of lips) to Bottom of Chin', '11a. (Left Side of Face to Left Side of eyes) to Centre of Left Pupil', '11b. (Right Side of Face to Right Side of eyes) to Centre of Right Pupil', '12a. (Left Side of Face to Left Side of Iris) to Other Side of Left eye', '12b. (Right Side of Face to Right Side of Iris) to Other Side of Right eye', '13a. (Left Side of Face to Left Side of Iris) to Centre of Face', '13b. (Right Side of Face to Right Side of Iris) to Centre of Face', '14a. (Left Side of Face to Side of Left eye) to the Right eye', '14b. (Right Side of Face to Side of Right eye) to the Left eye', '15a. (Left Side of Face to Centre of Face) to the Right eye', '15b. (Right Side of Face to Centre of Face) to the Left eye', '16a. (Left Side of Face to Inside Right eye) to Right side of face', '16b. (Right Side of Face to Inside Left eye) to Left side of face', '17a. (Left Side of Eye to Left Flair of Nose) to Right Flair of Nose', '17b. (Right Side of Eye to Right Flair of Nose) to Left Flair of Nose', '18a. (Centre of Left Pupil to Top of Nose Flair) to bottom of nose', '18b. (Centre of Right Pupil to Top of Nose Flair) to bottom of nose', '19a. (Centre of Left Pupil to Centre of Left Nostril) to Centre of Right Nostril', '19b. (Centre of Right Pupil to Centre of Right Nostrils) to Centre of Left Nostril', '20a. (Inside of Left Eye to Left Nose Bridge) to Right Nose Bridge', '20b. (Inside of Right Eye to Right Nose Bridge) to Left Nose Bridge', '21a. (Left Side of Mouth to Cupid’s Left bow point of lips) to Cupid’s Right bow point of lips', '21b. (Right Side of Mouth to Cupid’s Right bow point of lips) to Cupid’s Left bow point of lips', '22. (Bottom of nose to Top of lips) to Bottom of lips', '23. Head Height to Head Width']
    #plt.show()
    pcts = []
    n = 3
    for (name,z,graph) in zip(varNames, variables,graphs):
        cal = (max(z[0],z[1])-min(z[0],z[1]))/(max(z[2],z[3])-min(z[2],z[3]))
        #100% if difrence of cal and GoldenRatio = 0
        #0% if diffrence of cal and GoldenRatio = GoldenRatio
        #raw = max(cal,GoldenRatio)-min(cal,GoldenRatio)
        pct = 100* math.exp(-(abs(cal-GoldenRatio)/GoldenRatio))#Mean Percentge Error
        print(name+": "+str(pct)+"%")
        pcts.append(pct)
        worksheet.insert_image(index,n,resize(graph,224))
        worksheet.write(index,n+1,str(pct)+"%",data_format)
        n+=2
    overall = avg(*pcts)
    worksheet.write(index,2,str(overall)+"%",data_format)
    pprint("Overall Rating: "+str(overall)+"%")
    MassPct.append(overall)
    


pprint("Preparing xlsxWriter...")
workbook = xlsxwriter.Workbook('FinalOutput.xlsx')
# The workbook object is then used to add new
# worksheet via the add_worksheet() method.
global worksheet
global data_format
worksheet = workbook.add_worksheet()
worksheet.set_default_row(hide_unused_rows=True)
worksheet.set_cols.setdefault(1)

# Use the worksheet object to write
# data via the write() method.
heading_format = workbook.add_format({
    'bold': 1,
    'align': 'center',
    'valign': 'vcenter'})
data_format = workbook.add_format({
    'align': 'left',
    'valign': 'vcenter'})
data_format.set_text_wrap()
worksheet.merge_range('A1:A2', 'Name', heading_format)
worksheet.merge_range('B1:B2', 'Keypoint Scan', heading_format)
worksheet.merge_range('C1:C2', 'Overall Percentage', heading_format)
'''worksheet.merge_range('D1:D2', 'Original AI Scan', heading_format)
worksheet.merge_range('E1:E2', 'Scan After Refinements', heading_format)'''
worksheet.set_column('A:A', 36)
worksheet.set_column('B:B', 32)
'''worksheet.set_column('C:C', 32)
worksheet.set_column('D:D', 32)
worksheet.set_column('E:E', 32)'''
data_format = workbook.add_format({
    'align': 'center',
    'valign': 'vcenter'})
data_format.set_text_wrap()
n = 3
varNames = ['1. (Eyes to Nose Flair) to Nose Base', '2. (Eyes to Nostril top) to Centre of lips', '3. (Eyes to Nose base) to Bottom of lips', '4. (Eyes to Centre of lips) to Bottom of Chin', '5. (Nose Flair to Bottom of Lips) to Bottom of Chin', '6. (Nose Flair to Top of Lips) to Bottom of Chin', '7. (Top of Lips to Bottom of lips) to Bottom of Chin', '8. (Top of Lips to Centre of lips) to Bottom of Lips', '9. (Top of Eyebrows to Top of eyes) to Bottom of eyes', '10. (Top of Eyebrows to Top of lips) to Bottom of Chin', '11a. (Left Side of Face to Left Side of eyes) to Centre of Left Pupil', '11b. (Right Side of Face to Right Side of eyes) to Centre of Right Pupil', '12a. (Left Side of Face to Left Side of Iris) to Other Side of Left eye', '12b. (Right Side of Face to Right Side of Iris) to Other Side of Right eye', '13a. (Left Side of Face to Left Side of Iris) to Centre of Face', '13b. (Right Side of Face to Right Side of Iris) to Centre of Face', '14a. (Left Side of Face to Side of Left eye) to the Right eye', '14b. (Right Side of Face to Side of Right eye) to the Left eye', '15a. (Left Side of Face to Centre of Face) to the Right eye', '15b. (Right Side of Face to Centre of Face) to the Left eye', '16a. (Left Side of Face to Inside Right eye) to Right side of face', '16b. (Right Side of Face to Inside Left eye) to Left side of face', '17a. (Left Side of Eye to Left Flair of Nose) to Right Flair of Nose', '17b. (Right Side of Eye to Right Flair of Nose) to Left Flair of Nose', '18a. (Centre of Left Pupil to Top of Nose Flair) to bottom of nose', '18b. (Centre of Right Pupil to Top of Nose Flair) to bottom of nose', '19a. (Centre of Left Pupil to Centre of Left Nostril) to Centre of Right Nostril', '19b. (Centre of Right Pupil to Centre of Right Nostrils) to Centre of Left Nostril', '20a. (Inside of Left Eye to Left Nose Bridge) to Right Nose Bridge', '20b. (Inside of Right Eye to Right Nose Bridge) to Left Nose Bridge', '21a. (Left Side of Mouth to Cupid’s Left bow point of lips) to Cupid’s Right bow point of lips', '21b. (Right Side of Mouth to Cupid’s Right bow point of lips) to Cupid’s Left bow point of lips', '22. (Bottom of nose to Top of lips) to Bottom of lips', '23. Head Height to Head Width']
#count = 1
for vari in varNames:
    worksheet.merge_range(0,n,0,n+1, vari, heading_format)
    worksheet.set_column(1,n, 32)
    #worksheet.set_column(1,n+1, 8)
    worksheet.write(1,n,'Scan', heading_format)
    worksheet.write(1,n+1,'Rating', heading_format)
    #count += 1
    n += 2
worksheet.set_column(1,n-1, 32)
worksheet.set_column(1,n, 32)
worksheet.write(0,n,'Graph of Overall', heading_format)



pprint("Loading Excel Data...")
k=0
for a in range(3,m_col+1):
    #print("| ",end = "")
    if sheet.cell(row = a, column = 1).value == None:
        #print(sheet.cell(row = a, column = 1).value)
        continue
    name = str(sheet.cell(row = a, column = 1).value)
    pprint("Loading " + name)
    worksheet.set_row(a, 180)
    worksheet.write(a,0,name, data_format)
    fig, ax = plt.subplots()
    i = plt.imread(im[a-3])
    ax.imshow(i, extent=[0, 223.5, 223.5, 0])
    xPlots = []
    yPlots = []
    data = []
    for b in range(1,m_col+1):
        val = sheet.cell(row = a, column = b).value
        if val == None:
            continue
        try:
            num = float(val)
            if num.is_integer():
                num = int(num)
            if b%2 == 0:
                xPlots.append(num)
            else:
                yPlots.append(num)
        except ValueError:
            pass
    for (x,y) in zip(xPlots, yPlots):
        data.append({'x':x,'y':y})
    plt.scatter(xPlots,yPlots)
    if not os.path.exists("./xlsxReaderCache"):
        os.makedirs("./xlsxReaderCache")
    img = "./xlsxReaderCache/"+str(time.time()).replace(".","")+".png"
    plt.savefig(img)
    plt.close()
    worksheet.insert_image(a,1, resize(img,224))
    Calculate(i, str(sheet.cell(row = a, column = 1).value), data, a)
    k=a
worksheet.merge_range(1,n,k,n, "")



fig, ax = plt.subplots(figsize=(len(MassPct)*0.375, 6))


count = range(1,len(MassPct)+1)
a = []
for items in count:
    a.append(str(items))

ax.bar(a,MassPct, align='center', width = 0.75)


print(avg(*MassPct))
plt.axhline(avg(*MassPct), color='orangered',linestyle = '-')
label = ax.text(len(MassPct)+1, avg(*MassPct), "Average Percentage", color='orangered', fontsize=10,va="bottom", ha="center")
ax.set_ylabel('Overall Percentage')
ax.set_xlabel('Index Number')
ax.set_title('Class\'s Overall Percentage')
ax.set_ylim([0, 100])
plt.subplots_adjust(right=0.875)
g = "./xlsxReaderCache/"+str(time.time()).replace(".","")+".png"
plt.savefig(g)
plt.show()
plt.close()

worksheet.insert_image(1,n,g)

pprint("Closing Workbook...")
wb.close()
workbook.close()
